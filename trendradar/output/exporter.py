"""
TrendRadar SEO — output/exporter.py
=====================================
Exporta los resultados a CSV y JSON para análisis posterior.

Funcionalidades:
  - export_csv()   → archivo CSV con todos los scored topics y sus noticias
  - export_json()  → archivo JSON con estructura jerárquica completa
  - Timestamp en nombre del archivo para evitar sobreescrituras
  - Crea el directorio de exportación si no existe
"""

import csv
import json
import logging
import os
import unicodedata
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

from trendradar.config import OUTPUT_DIR, CSV_FILENAME, JSON_FILENAME

logger = logging.getLogger(__name__)


def export_csv(
    scored_topics: list,
    news_by_topic: dict = None,
    output_dir:    str  = OUTPUT_DIR,
    filename:      str  = CSV_FILENAME,
) -> str:
    """
    Exporta los ScoredTopics y sus noticias a un archivo CSV.

    Columnas:
      rank, keyword, geo, timeframe, trend_score, discover_tier,
      interest_peak, interest_now, interest_avg, growth_pct, is_breakout,
      is_fresh_trend, is_accelerating, is_sustained,
      recommendation, related_queries,
      news_count, news_fresh_count,
      news_1_title, news_1_url, news_1_source, news_1_published,
      news_2_title, ... (hasta 5 noticias)

    Returns:
        Ruta absoluta del archivo creado.
    """
    news_by_topic = news_by_topic or {}
    path = _make_output_path(output_dir, filename, ext="csv")

    rows = []
    for rank, s in enumerate(scored_topics, 1):
        news_list   = news_by_topic.get(s.keyword, [])
        news_count  = len(news_list)
        fresh_count = sum(1 for a in news_list if a.is_fresh)

        row = {
            "rank":            rank,
            "keyword":         s.keyword,
            "geo":             s.geo,
            "timeframe":       s.timeframe_label,
            "trend_score":     round(s.trend_score, 1),
            "discover_tier":   s.discover_tier,
            "interest_peak":   s.interest_peak,
            "interest_now":    s.interest_now,
            "interest_avg":    round(s.interest_avg, 2),
            "growth_pct":      round(s.growth_pct, 1),
            "is_breakout":     s.is_breakout,
            "is_fresh_trend":  s.is_fresh_trend,
            "is_accelerating": s.is_accelerating,
            "is_sustained":    s.is_sustained,
            "recommendation":  s.recommendation,
            "related_queries": " | ".join(s.related_queries),
            "news_count":      news_count,
            "news_fresh_count":fresh_count,
        }

        # Agregar columnas de noticias (hasta 5)
        for n_idx in range(1, 6):
            prefix = f"news_{n_idx}_"
            if n_idx <= len(news_list):
                art = news_list[n_idx - 1]
                pub = art.published.isoformat() if art.published else ""
                row[f"{prefix}title"]     = art.title
                row[f"{prefix}url"]       = art.url
                row[f"{prefix}source"]    = art.source
                row[f"{prefix}published"] = pub
                row[f"{prefix}fresh"]     = art.is_fresh
                row[f"{prefix}relevance"] = round(art.relevance_score, 2)
            else:
                row[f"{prefix}title"]     = ""
                row[f"{prefix}url"]       = ""
                row[f"{prefix}source"]    = ""
                row[f"{prefix}published"] = ""
                row[f"{prefix}fresh"]     = ""
                row[f"{prefix}relevance"] = ""

        rows.append(row)

    if not rows:
        logger.warning("export_csv: no hay datos para exportar.")
        return ""

    fieldnames = list(rows[0].keys())
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)

    logger.info("CSV exportado → %s (%d filas)", path, len(rows))
    return str(path)


def export_json(
    scored_topics: list,
    news_by_topic: dict = None,
    output_dir:    str  = OUTPUT_DIR,
    filename:      str  = JSON_FILENAME,
    metadata:      dict = None,
) -> str:
    """
    Exporta los ScoredTopics y sus noticias a un archivo JSON.

    Estructura:
    {
      "metadata": {
        "generated_at": "...",
        "geo": "CO",
        "timeframe": "Último día",
        "total_topics": 20
      },
      "topics": [
        {
          "rank": 1,
          "keyword": "...",
          "trend_score": 82.5,
          "discover": { "tier": "ALTO", "signals": {...} },
          "metrics": { "peak": 100, "now": 87, ... },
          "recommendation": "...",
          "related_queries": [...],
          "news": [
            { "title": "...", "url": "...", "source": "...", "published": "...", "is_fresh": true }
          ]
        }
      ]
    }

    Returns:
        Ruta absoluta del archivo creado.
    """
    news_by_topic = news_by_topic or {}
    path = _make_output_path(output_dir, filename, ext="json")

    output = {
        "metadata": {
            "generated_at": datetime.now(tz=timezone.utc).isoformat(),
            "total_topics": len(scored_topics),
            **(metadata or {}),
        },
        "topics": [],
    }

    for rank, s in enumerate(scored_topics, 1):
        news_list = news_by_topic.get(s.keyword, [])

        topic_data = {
            "rank":    rank,
            "keyword": s.keyword,
            "geo":     s.geo,
            "timeframe": s.timeframe_label,
            "trend_score": round(s.trend_score, 1),
            "discover": {
                "tier":           s.discover_tier,
                "has_breakout":   s.has_breakout,
                "is_fresh_trend": s.is_fresh_trend,
                "is_accelerating":s.is_accelerating,
                "is_sustained":   s.is_sustained,
            },
            "score_components": {
                "peak":     round(s.score_peak, 1),
                "velocity": round(s.score_velocity, 1),
                "curve":    round(s.score_curve, 1),
                "momentum": round(s.score_momentum, 1),
            },
            "metrics": {
                "interest_peak": s.interest_peak,
                "interest_now":  s.interest_now,
                "interest_avg":  round(s.interest_avg, 2),
                "growth_pct":    round(s.growth_pct, 1),
                "is_breakout":   s.is_breakout,
            },
            "recommendation":  s.recommendation,
            "related_queries": s.related_queries,
            "news": [
                {
                    "title":     art.title,
                    "url":       art.url,
                    "source":    art.source,
                    "published": art.published.isoformat() if art.published else None,
                    "is_fresh":  art.is_fresh,
                    "relevance": round(art.relevance_score, 2),
                    "snippet":   art.snippet,
                }
                for art in news_list
            ],
        }
        output["topics"].append(topic_data)

    with open(path, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    logger.info("JSON exportado → %s", path)
    return str(path)


def export_excel(
    scored_topics: list,
    news_by_topic: dict,
    ai_data_by_topic: dict,
    output_dir: str = OUTPUT_DIR,
    template_path: str = "PLANTILLA EXCEL.xlsx",
) -> str:
    """
    Exporta los resultados a la plantilla Excel del usuario.
    """
    try:
        from openpyxl import load_workbook
        import shutil
    except ImportError:
        logger.error("openpyxl no está instalado. No se puede exportar a Excel.")
        return ""

    if not os.path.exists(template_path):
        logger.error(f"No se encontró la plantilla: {template_path}")
        return ""

    path = _make_output_path(output_dir, "analisis_seo.xlsx", ext="xlsx")
    
    # Copiar plantilla
    shutil.copy2(template_path, path)
    
    wb = load_workbook(path)
    
    # 1. Hoja: Inteligencia de Tendencia
    ws_intel = _get_sheet(wb, "Inteligencia de Tendencia")
    if ws_intel:
        # Filas empiezan en 3 (1 = título, 2 = headers)
        for i, s in enumerate(scored_topics, start=3):
            # 'TrendScore', 'Keyword / Tema', 'Tier Discover', 'Volumen / Tráfico', 'Breakout', 'Crecimiento %'
            ws_intel.cell(row=i, column=1, value=round(s.trend_score, 1))
            ws_intel.cell(row=i, column=2, value=s.keyword)
            ws_intel.cell(row=i, column=3, value=s.discover_tier)
            # El tráfico aproximado está en el raw_topic, asumiendo que lo podemos rescatar, o usar interest_peak
            # ya que no guardamos el approx_traffic exacto en ScoredTopic, usamos score o si lo agregamos después
            traffic = getattr(s.topic, "approx_traffic", "")
            ws_intel.cell(row=i, column=4, value=traffic)
            ws_intel.cell(row=i, column=5, value="Sí" if s.has_breakout else "No")
            ws_intel.cell(row=i, column=6, value=f"+{round(s.growth_pct, 1)}%")

    # 2. Hoja: Planificación Editorial
    ws_plan = _get_sheet(wb, "Planificación Editorial")
    if ws_plan:
        for i, s in enumerate(scored_topics, start=3):
            # 'Ángulo / Enfoque', 'Keyword Principal (SEO)', 'Entidades (NLP)', 'Clickbait Ético'
            ai_data = ai_data_by_topic.get(s.keyword, {})
            ws_plan.cell(row=i, column=1, value=ai_data.get("angulo", ""))
            ws_plan.cell(row=i, column=2, value=s.keyword)
            ws_plan.cell(row=i, column=3, value=ai_data.get("entidades", ""))
            ws_plan.cell(row=i, column=4, value=ai_data.get("clickbait", ""))

    # 3. Hoja: Ejecución y Tracking
    ws_track = _get_sheet(wb, "Ejecución y Tracking")
    if ws_track:
        # Limpiar headers de ejemplo si existen y poner los correctos
        ws_track.cell(row=2, column=1, value="Estado")
        ws_track.cell(row=2, column=2, value="Responsable")
        ws_track.cell(row=2, column=3, value="Link URL")
        ws_track.cell(row=2, column=4, value="Tráfico 24h")
        
        # Limpiar fila 3 si tenía datos de Markdown
        for col in range(1, 5):
            ws_track.cell(row=3, column=col, value="")

    # 4. Hoja dinámica: Expansión Long Tail (KeySearch)
    ws_longtail = wb.create_sheet("Expansión Long Tail")
    ws_longtail.append(["Tendencia Padre", "Tipo de Variación", "Keyword Long Tail / Pregunta"])
    row_lt = 2
    for s in scored_topics:
        padre = s.keyword
        for q in getattr(s, "paa_questions", []):
            ws_longtail.cell(row=row_lt, column=1, value=padre)
            ws_longtail.cell(row=row_lt, column=2, value="PAA (Pregunta Frecuente)")
            ws_longtail.cell(row=row_lt, column=3, value=q)
            row_lt += 1
            
        for q in getattr(s, "autocomplete_suggestions", []):
            ws_longtail.cell(row=row_lt, column=1, value=padre)
            ws_longtail.cell(row=row_lt, column=2, value="Autocompletado")
            ws_longtail.cell(row=row_lt, column=3, value=q)
            row_lt += 1
            
        for q in getattr(s, "related_queries_longtail", []):
            ws_longtail.cell(row=row_lt, column=1, value=padre)
            ws_longtail.cell(row=row_lt, column=2, value="Búsqueda Relacionada")
            ws_longtail.cell(row=row_lt, column=3, value=q)
            row_lt += 1

    wb.save(path)
    logger.info("Excel exportado → %s", path)
    return str(path)


# ──────────────────────────────────────────────
# Helpers
# ──────────────────────────────────────────────
def _make_output_path(output_dir: str, filename: str, ext: str) -> Path:
    """
    Genera la ruta del archivo de salida con timestamp.
    Crea el directorio si no existe.

    Ejemplo: exports/trendradar_export_20250613_014500.csv
    """
    dir_path = Path(output_dir)
    dir_path.mkdir(parents=True, exist_ok=True)

    # Insertar timestamp antes de la extensión
    stem  = Path(filename).stem
    ts    = datetime.now().strftime("%Y%m%d_%H%M%S")
    fname = f"{stem}_{ts}.{ext}"

    return dir_path / fname


def _get_sheet(workbook, expected_name: str):
    """Busca una hoja ignorando espacios accidentales y acentos."""
    expected = _normalize_sheet_name(expected_name)
    for name in workbook.sheetnames:
        if _normalize_sheet_name(name) == expected:
            return workbook[name]
    logger.warning("No se encontro la hoja '%s' en la plantilla.", expected_name)
    return None


def _normalize_sheet_name(name: str) -> str:
    text = unicodedata.normalize("NFKD", name.strip().lower())
    return "".join(ch for ch in text if not unicodedata.combining(ch))
